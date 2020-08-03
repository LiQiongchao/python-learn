import os
import openpyxl
import datetime
import time
start_time = time.time()
file_path=r"D:\聚类"
file_name_source = "数据.xlsx"
file_name_target = "结果.xlsx"


wb_s = openpyxl.load_workbook(os.path.join(file_path,file_name_source)) #导入原表格
ws_s = wb_s["详表"]
ws_city=wb_s["区县"]

wb_t = openpyxl.load_workbook(os.path.join(file_path,file_name_target))
ws_t = wb_t["详表"]

def getdistance(lata, lona, latb, lonb): #计算距离函数
    x=((lata-latb)*111000)**2
    y=((lona-lonb)*111000)**2
    d=(x+y)**(1/2)
    return(d)


#找到每个区县对应的起点和重点位置，用于循环
city=[]
cityadd={}
for i in range(1,ws_s.max_row+1):
    city.append(ws_s.cell(row=i,column=5).value)
print(len(city))

for i in range(1,ws_city.max_row+1):  #区县表内容
    key=ws_city.cell(row=i,column=1).value
    value = city.index(key)+1 #city的起始位置
    cityadd.setdefault(key,[]).append(value)
    value1= city.count(key)+value  #city的结束位置
    cityadd.setdefault(key,[]).append(value1)
print(cityadd)


ws_t_row=1 #从第一行开始写入表格
ws_t_team=1

for i in range(2,ws_s.max_row+1): #遍历源表格
    
    
    if ws_s.cell(row=i,column=10).value == 1: #如果已计算，跳过此次循环
        continue

    team=[] #聚类目标
    teamrow=[] 
    teamcount=1 #目标队列里面循环计算次数
    ws_s.cell(row=i,column=10).value =1
    c = ws_s.cell(row=i,column=5).value#获取区县信息
    qidian = cityadd[c][0]
    zhongdian=cityadd[c][1]

    for j in range(qidian,zhongdian): #循环I行的CELL
        if j == i:
            continue
        lata = ws_s.cell(row=i,column=2).value
        lona = ws_s.cell(row=i,column=3).value
        latb = ws_s.cell(row=j,column=2).value
        lonb = ws_s.cell(row=j,column=3).value
        d = getdistance(lata,lona,latb,lonb)
        if d < 50: #聚类距离
            team.append(ws_s.cell(row=i,column=1).value)
            team.append(ws_s.cell(row=j,column=1).value)
            teamrow.append(i)
            teamrow.append(j)
            ws_s.cell(row=j,column=10).value=1 #标识已计算

    while teamcount < len(team): #循环team里面的其他值
        
        for k in range(qidian,zhongdian):
            if k == i or k == teamrow[teamcount]:
                continue
            lata = ws_s.cell(row=teamrow[teamcount],column=2).value
            lona = ws_s.cell(row=teamrow[teamcount],column=3).value
            latb = ws_s.cell(row=k,column=2).value
            lonb = ws_s.cell(row=k,column=3).value
            d = getdistance(lata,lona,latb,lonb)

            if d <50 and k not in teamrow:
                team.append(ws_s.cell(row=k,column=1).value)
                teamrow.append(k)
                ws_s.cell(row=k,column=10).value =1

        teamcount +=1

    if teamcount >= 3:  #把TEAM里面的值写入聚类目标表格
        for l in range(0,teamcount):
            ws_t.cell(row=ws_t_row,column=1).value = ws_t_team
            ws_t.cell(row=ws_t_row,column=2).value = team[l]
            ws_t_row +=1
        ws_t_team += 1



wb_t.save("聚类.xlsx")
print(f'完成，用时为{time.time()-start_time}秒')
                            
                            
            
        
    
        
        



