"""
计算 Excel 数据

@Author: QiongchaoLi
@Date: 2020/8/29 10:08
"""
import csv
import os
import time

import openpyxl

# 根文件夹
BASE_DIR = 'D:\\weiyunSyncFolder\\SyncFolder-653028346\\temp\\zhong'

# 要处理的所有的CSV文件的文件夹名
SOURCE_DIR = 'calculate'
# 输出结果的文件名
RESULT_DIR = 'calculateResult'
RESULT_FILE_NAME = 'result-统一时段小区级报表-天-' + time.strftime("%Y-%m-%d", time.localtime()) + '.xlsx'

# 结果集在总结果中的下标（即要计算的列的位置，如‘上行PRB平均利用率(%)’的位置是4，从0开始。）
result_indexes = (4, 13, 16, 19, 20, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 73, 74, 75, 79)
ave_indexes = [1, 2, 3, 4, 5, 6, 10, 11, 14, 15, 16, 17, 18]
max_indexes = [12, 13, 19]
sum_indexes = [7, 8, 9, 20]
# RPC最大连接数的列数
RPC_connect_max_count_col = 36

# RPC最大连接数 取 top几，默认取top3
top_num_of_RPC_connect_max_count = 3

# 结果集的头信息
result_header = []
# 标识是否计算过表头，只提取一次
hasHeader = False
# {ecgi:[], ecgi2:[]}
result_rows = {}
# 统计ecgi的个数，用于求平均
# {ecgi1, 1, ecgi2, 3}
ecgi_cal = {}
# 找出每个 ECGI 的3个最大值
ecgi_max_map = {}


# 读取所有的csv文件
def read_files(fil_dir):
    if not fil_dir:
        print('未指定文件目录')
        raise ValueError('无效目录地址')
    file_list = []
    for root, dirs, files in os.walk(fil_dir):
        for file in files:
            if file.endswith('.csv'):
                file_list.append(file)
    return file_list


# 计算csv文件
def save_date_to_excel(result_header, result_rows, ecgi_cal, ecgi_max_map):
    # 计算平均值，总和，等值
    wb = openpyxl.Workbook()
    ws1 = wb.active
    # 写入头标题
    for i in range(len(result_header)):
        ws1.cell(row=1, column=i + 1, value=result_header[i])
    # 写入行数据
    row_num = 2
    for ecgi_key in result_rows.keys():
        result_row = result_rows.get(ecgi_key)
        ecgi = result_row[0]
        for col in range(len(result_header)):
            cell_val = result_row[col]
            if col in ave_indexes:
                cell_val = cell_val / ecgi_cal.get(ecgi)
            ws1.cell(row_num, col + 1, cell_val)
        row_num += 1

    print(RESULT_FILE_NAME, " 计算完成！")
    result_file = os.path.join(BASE_DIR, RESULT_DIR, RESULT_FILE_NAME)
    wb.save(result_file)
    print(result_file, ' 写入完成！')


# 计算单个文件
def cal_file(file_dir, file):
    # 拼接完全文件路径名
    fill_file = os.path.join(file_dir, file)

    # 读取一个 CSV 文件
    with open(fill_file, encoding='utf-8') as f:
        source_list = list(csv.reader(f))
        # 标识使用的是全局变量，否则使用创建一个局部变量
        global hasHeader
        if not hasHeader:
            # 提取头的名称（即第一行的信息）
            header = list(source_list[0])
            # 提取结果的文件头名称
            for count in result_indexes:
                result_header.append(header[count])
            hasHeader = True
        # 循环取每一行信息（从第二行开始取）
        for i in range(1, len(source_list)):
            row = source_list[i]

            ecgi = row[4]
            ecgi_cal[ecgi] = ecgi_cal.setdefault(ecgi, 0) + 1

            # 取出当前需要计算的数据
            # [ecgi, 上行PRB，下午PRB。。。]
            row_for_max = [ecgi]

            # 初始化列表，长度为result_indexes的长度，并赋初始值为 0.0
            init_row = [0.0 for i in range(len(result_indexes))]
            init_row.insert(0, ecgi)
            result_row = result_rows.setdefault(ecgi, init_row)
            # 根据ECGI计算平均
            for j in range(1, len(result_indexes)):
                if j not in ave_indexes:
                    continue
                # 取原始值，用于计算是否是三个最大值
                cell_val = row[result_indexes[j]]
                if not cell_val:
                    cell_val = 0
                float_val = float(cell_val)
                row_for_max.append(float_val)
                result_row[j] += float_val

            result_rows[ecgi] = result_row

    # return result_rows, ecgi_cal, ecgi_max_map


def cal_csv(file_dir):
    # 获取目录下的所有文件名
    files = read_files(file_dir)
    for file in files:
        # result_rows, ecgi_cal, ecgi_max_map = cal_file(file_dir, file)
        cal_file(file_dir, file)
        print(file, ' 统计完成！')

    # 保存结果到excel
    save_date_to_excel(result_header, result_rows, ecgi_cal, ecgi_max_map)


def main():
    start_time = time.time()
    cal_csv(os.path.join(BASE_DIR, SOURCE_DIR))
    print(f'完成，用时: {time.time() - start_time} 秒')


if __name__ == '__main__':
    main()
