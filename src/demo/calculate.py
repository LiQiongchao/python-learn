"""
计算 Excel 数据

@Author: QiongchaoLi
@Date: 2020/8/29 10:08
"""
import csv
import os
import time
from pathlib import Path

import openpyxl

# 根文件夹
BASE_DIR = 'D:\\weiyunSyncFolder\\SyncFolder-653028346\\temp\\zhong'

# 要处理的所有的CSV文件的文件夹名
SOURCE_DIR = 'calculate'
# 输出结果的文件名
RESULT_DIR = 'calculateResult'

# 结果集在总结果中的下标（即要计算的列的位置，如‘上行PRB平均利用率(%)’的位置是4，从0开始。）
result_indexes = (4, 13, 16, 19, 20, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 73, 74, 75, 79)
ave_indexes = [1, 2, 3, 4, 5, 6, 10, 11, 14, 15, 16, 17, 18]
max_indexes = [12, 13, 19]
sum_indexes = [7, 8, 9, 20]
# RPC最大连接数的列数
RPC_connect_max_count = 36


# 读取所有的csv文件
def read_files(fil_dir):
    if not fil_dir:
        print('未指定文件目录')
        raise ValueError('无效目录地址')
    file_list = []
    for root, dirs, files in os.walk(fil_dir):
        for file in files:
            if file.endswith('.csv'):
                file_list.append(file)
    return file_list


# 计算csv文件
def save_date_to_excel(file, result_header, result_rows, ecgi_cal, ecgi_max_map):
    # 计算平均值，总和，等值
    wb = openpyxl.Workbook()
    ws1 = wb.active
    # 写入头标题
    for i in range(len(result_header)):
        ws1.cell(row=1, column=i + 1, value=result_header[i])
    # 写入行数据
    row_num = 2
    for ecgi_key in result_rows.keys():
        result_row = result_rows.get(ecgi_key)
        ecgi = result_row[0]
        for col in range(len(result_header)):
            cell_val = result_row[col]
            if col in ave_indexes:
                cell_val = cell_val / ecgi_cal.get(ecgi)
            ws1.cell(row_num, col + 1, cell_val)
        row_num += 1
    # 写入每个ecgi的三个最大值
    ws2 = wb.create_sheet('Sheet1')
    # 写入头标题
    for i in range(len(result_header)):
        ws2.cell(row=1, column=i + 1, value=result_header[i])
    row_num = 2
    for ecgi_key in ecgi_max_map.keys():
        max_third = ecgi_max_map.get(ecgi_key)
        for i in range(len(max_third)):
            max_vals = max_third[i]
            for col in range(len(result_header)):
                ws2.cell(row_num, col + 1, max_vals[col])
        row_num += 1
    file_name = file.rsplit(".")[0] + '.xlsx'
    print(file_name, " 计算完成！")
    wb.save(os.path.join(BASE_DIR, RESULT_DIR, file_name))


def cal_file(file_dir, file):
    # 拼接完全文件路径名
    fill_file = os.path.join(file_dir, file)
    # 结果集的头信息
    result_header = []
    # {ecgi:[], ecgi2:[]}
    result_rows = {}
    # 统计ecgi的个数，用于求平均
    # {ecgi1, 1, ecgi2, 3}
    ecgi_cal = {}
    # 找出每个 ECGI 的3个最大值
    ecgi_max_map = {}
    # 读取一个 CSV 文件
    with open(fill_file, encoding='utf-8') as f:
        source_list = list(csv.reader(f))
        # 提取头的名称（即第一行的信息）
        header = list(source_list[0])
        # 提取结果的文件头名称
        for count in result_indexes:
            result_header.append(header[count])
        # 循环取每一行信息（从第二行开始取）
        for i in range(1, len(source_list)):
            row = source_list[i]

            ecgi = row[4]
            ecgi_cal[ecgi] = ecgi_cal.setdefault(ecgi, 0) + 1

            # 取出当前需要计算的数据
            # [ecgi, 上行PRB，下午PRB。。。]
            row_for_max = [ecgi]

            # 初始化列表，长度为result_indexes的长度，并赋初始值为 0.0
            init_row = [0.0 for i in range(len(result_indexes))]
            init_row.insert(0, ecgi)
            result_row = result_rows.setdefault(ecgi, init_row)
            # 根据ECGI计算平均，总和，最大等信息
            temp_12 = result_row[12]
            temp_13 = result_row[13]
            temp_19 = result_row[19]
            for j in range(1, len(result_indexes)):
                # 取原始值，用于计算是否是三个最大值
                cell_val = row[result_indexes[j]]
                if not cell_val:
                    cell_val = 0
                float_val = float(cell_val)
                row_for_max.append(float_val)
                result_row[j] += float_val
            # 计算并放入最大值
            result_row[12] = max(temp_12, float(row[result_indexes[12]] if row[result_indexes[12]] else 0))
            result_row[13] = max(temp_13, float(row[result_indexes[13]] if row[result_indexes[13]] else 0))
            result_row[19] = max(temp_19, float(row[result_indexes[19]] if row[result_indexes[19]] else 0))
            result_rows[ecgi] = result_row

            # 计算关于最大值的问题
            max_third = list()
            if ecgi in ecgi_max_map.keys():
                max_third = list(ecgi_max_map.get(ecgi))

            if len(max_third) < 3:
                max_third.append(row_for_max)
                # 按 上行PRB平均利用率(%) 升序排序
                max_third.sort(key=lambda elem: elem[12])
                ecgi_max_map[ecgi] = max_third
            else:
                for i in range(3):
                    temp = max_third[i]
                    if temp[12] < row_for_max[12]:
                        max_third[i] = row_for_max
                        max_third.sort(key=lambda elem: elem[12])
                        ecgi_max_map[ecgi] = max_third
                        break

    # 保存结果到excel
    save_date_to_excel(file, result_header, result_rows, ecgi_cal, ecgi_max_map)
    # return result_rows, ecgi_cal, ecgi_max_map


def cal_csv(file_dir):
    # 获取目录下的所有文件名
    files = read_files(file_dir)
    for file in files:
        # result_rows, ecgi_cal, ecgi_max_map = cal_file(file_dir, file)
        cal_file(file_dir, file)


def main():
    start_time = time.time()
    cal_csv(os.path.join(BASE_DIR, SOURCE_DIR))
    print(f'完成，用时: {time.time() - start_time} 秒')


if __name__ == '__main__':
    main()
